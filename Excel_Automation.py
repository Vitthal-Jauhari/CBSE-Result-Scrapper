from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
import time
import browser_scrapper


def printslow(string, delay=0.02):
    string += '\n'
    for letter in string:
        print(letter, end='', flush=True)
        # time.sleep(0.02)


def Get_admit_card_ID(Name, Roll_no, Mother_name):
    try:
        if not Mother_name:
            raise ValueError("Mother's name is missing")
        AdmitCardID = Name[0] + Mother_name[0] + str(Roll_no)[-3:-1] + str(7072)
        return AdmitCardID
    except Exception as e:
        print('Error while compiling Admit Card:', e)
        print('Given Resources:', [Name, Roll_no, Mother_name])
        return None


def Setup_sheet(sheet):
    if sheet['A2'].value is None:
        sheet['A1'] = "No."
        sheet['B1'] = "Name"
        sheet['C1'] = "Board Roll No"
        sheet['D1'] = "Admit Card ID"
        sheet['E1'] = "Best 5 Percentage"
        sheet['F1'] = "Core 5 Percentage"
        bold_font = Font(bold=True, size=12)
        alignment = Alignment(horizontal='center', vertical='center')
        for cell in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1']:
            sheet[cell].font = bold_font
            sheet[cell].alignment = alignment


def Write_ID(sheet, row, Details):
    try:
        # wb = load_workbook(r'C:\Users\vitth\Documents\Class XII A details.xlsx')
        # sheet = wb['Sheet3']
        wb = Workbook(r"C:\Users\vitth\Desktop\Marks.xlsx")
        sheet = wb.active

        Name, Roll_no, AdmitCardID, best5, core5 = Details          # Unpacking the list
        sheet[f'A{row}'] = row-1
        sheet[f'B{row}'] = Name
        sheet[f'C{row}'] = Roll_no
        sheet[f'D{row}'] = AdmitCardID
        sheet[f'E{row}'] = best5
        sheet[f'F{row}'] = core5
        wb.save(r'"C:\Users\vitth\Desktop\Marks.xlsx"')
    except Exception as e:
        print("Error while writing: ", e)


def finalize_sheet(sheet):
    # Center alignment for every cell in the sheet
    alignment = Alignment(horizontal='center', vertical='center')
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = alignment

    # Auto-size columns based on the largest content in each column
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Adjusted width with padding
        sheet.column_dimensions[column].width = adjusted_width


def Read_Initials():
    wb = load_workbook(r'C:\Users\vitth\Documents\Class XII A details.xlsx')
    ws = wb["Sheet2"]
    # result_book = Workbook(r'')
    result_sheet = wb['Sheet3']
    # time.sleep(0.5)
    print('\nWorksheet loaded...\n')
    # time.sleep(1)

    try:
        current_row = 2                 # Start writing from the 2nd row after the headers
        for row in range(2,45):
            name        = ws['B'+str(row)].value
            mother_name = ws['J'+str(row)].value
            roll_no     = ws['C'+str(row)].value

            print('\n\nProcessing request: ', name)
            admit_card_ID = Get_admit_card_ID(name, roll_no, mother_name)
            if not admit_card_ID:
                continue                # If any parameter for the admit card is absent

            details = browser_scrapper.fetch_details(name, roll_no, admit_card_ID)
            # Write_ID(ws, current_row, details)

            Setup_sheet(wb['Sheet3'])
            result_sheet[f'A{current_row}'] = current_row -1
            result_sheet[f'B{current_row}'] = name
            result_sheet[f'C{current_row}'] = roll_no
            result_sheet[f'D{current_row}'] = admit_card_ID
            result_sheet[f'E{current_row}'] = details[-2]
            result_sheet[f'F{current_row}'] = details[-1]
            current_row += 1
            print('Details written: ', details)
            wb.save(r'C:\Users\vitth\Documents\Class XII A details.xlsx')

            print('Data saved.')
    except Exception as e:
        print('Error in main block: ', e)
    finally:
        finalize_sheet(result_sheet)                # Final formatting (spacing, alignment etc.)
    


def main():
    # time.sleep(1)
    print("\nScript starting...")

    # time.sleep(2)
    printslow('\nProgramming initialising...')

    Read_Initials()

if __name__ == '__main__':
    main()